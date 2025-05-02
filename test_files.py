import csv
import io
import zipfile
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
import tmp


def test_zip():
    with zipfile.ZipFile("resources/myzip.zip", "r") as test:
        print(test.infolist())


def test_read_csv():
    expected_data = [
        ['plu', 'shop_sap_id'],
        ['2057821', 'HD02'],
        ['3165030', 'HD02'],
        ['112397321', 'HD02'],
        ['4116701', 'HD02']
    ]
    archive_zip = "resources/myzip.zip"
    with zipfile.ZipFile(archive_zip, 'r') as zf:
        with zf.open('outpu1.csv', 'r') as csv_file:
            text_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.reader(text_file)
            actual_data = list(reader)

            assert actual_data == expected_data


def test_read_xlsx():
    expected_data = [
        ['new', 'old'],
        [1.0, 5.0], [2.0, 6.0], [3.0, 7.0], [4.0, 8.0]
    ]
    archive_zip = "resources/myzip.zip"
    with zipfile.ZipFile(archive_zip, 'r') as zf:
        with zf.open('item1.xlsx') as xlsx_file:
            workbook = load_workbook(filename=xlsx_file)
            sheet = workbook.active

            actual_data = []
            for row in sheet.iter_rows(values_only=True):
                actual_data.append(list(row))

            assert actual_data == expected_data


def test_read_pdf():
    with zipfile.ZipFile("resources/myzip.zip", 'r') as zf:
        with zf.open("pdf_sample.pdf") as pdf_file:
            pdf_reader = PdfReader(io.BytesIO(pdf_file.read()))
            expected_data = (
                '№ Column 1 Column 2 Column 3 Column 4 Column 5 \n'
                'Row 1 1 1 1 1 1 \n'
                'Row 2 2 2 2 2 2 \n'
                'Row 3 3 3 3 3 3 \n'
                'Row 4 4 4 4 4 4 \n'
                'Row 5 5 5 5 5 5 '
            )
            actual_data = ""
            for page in pdf_reader.pages:
                actual_data = page.extract_text()
            assert actual_data == expected_data
